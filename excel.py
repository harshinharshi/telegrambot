import os
from openpyxl import Workbook, load_workbook


def append_to_excel(data: dict, filename="finance_tracker.xlsx"):
    file_exists = os.path.exists(filename)
    
    if file_exists:
        # Load the existing workbook
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Ensure headers match dictionary keys
        headers = [cell.value for cell in sheet[1]]
        row = [data.get(header, "") for header in headers]
        sheet.append(row)
    else:
        # Create new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active

        # Write headers
        headers = list(data.keys())
        sheet.append(headers)

        # Write the first row of data
        row = list(data.values())
        sheet.append(row)

    workbook.save(filename)
    print(f"Data appended to {filename}")

# Example usage:
# data = {
#     "Date": "2025-06-01",
#     "Category": "Groceries",
#     "Amount": 150,
#     "Description": "Weekly supermarket run"
# }

# append_to_excel(data)
